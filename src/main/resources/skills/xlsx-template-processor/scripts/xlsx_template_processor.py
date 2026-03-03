"""
Excel模板处理器
基于提供的Excel模板和数据生成新的Excel文件
"""

import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def process_xlsx_template(
    template_path: str,
    data: Dict[str, Any],
    output_path: str,
    start_row_marker: Optional[str] = None
) -> str:
    """
    处理Excel模板并生成新的Excel文件
    
    Args:
        template_path: 模板文件路径
        data: 用于填充模板的数据字典
        output_path: 输出文件路径
        start_row_marker: 表格数据开始标记（可选）
        
    Returns:
        生成的Excel文件路径
    """
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 处理采购合同模板的特定逻辑
    if _is_purchase_contract_template(ws):
        _process_purchase_contract_template(ws, data)
    else:
        # 替换单元格中的占位符
        _replace_placeholders(ws, data)
        
        # 处理表格数据（如果存在）
        _process_table_data(ws, data, start_row_marker)
    
    wb.save(output_path)
    return output_path


def process_xlsx_template_with_context(
    template_path: str,
    data: Dict[str, Any],
    output_path: str,
    context_supplier_info: Optional[Dict[str, Any]] = None,
    start_row_marker: Optional[str] = None
) -> str:
    """
    处理Excel模板并生成新的Excel文件，支持上下文供应商信息
    
    Args:
        template_path: 模板文件路径
        data: 用于填充模板的数据字典
        output_path: 输出文件路径
        context_supplier_info: 从用户上下文获取的供应商信息（可选）
        start_row_marker: 表格数据开始标记（可选）
        
    Returns:
        生成的Excel文件路径
    """
    # 如果提供了上下文供应商信息，将其合并到data中
    merged_data = data.copy()
    if 'supplier' not in merged_data:
        merged_data['supplier'] = {}
    
    if context_supplier_info and isinstance(context_supplier_info, dict):
        merged_data['supplier'].update(context_supplier_info)
    
    return process_xlsx_template(template_path, merged_data, output_path, start_row_marker)


def _is_purchase_contract_template(ws) -> bool:
    """
    检测是否为采购合同模板
    """
    # 检查特定的关键标识
    cell_a2 = ws['A2'].value if ws['A2'].value else ""
    cell_a3 = ws['A3'].value if ws['A3'].value else ""
    
    # 如果包含"采购合同"字样，则认为是采购合同模板
    return "采购合同" in str(cell_a2) or "采购合同" in str(cell_a3) or "合同" in str(cell_a2)


def _process_purchase_contract_template(ws, data: Dict[str, Any]):
    """
    专门处理采购合同模板的逻辑
    根据实际模板结构进行适配
    """
    # 设置合同编号 - 根据实际模板位置 F2
    if 'contract_number' in data and data['contract_number']:
        _safe_write_cell(ws, 'F2', data['contract_number'])
    
    # 设置签订地点 - 根据实际模板位置（从分析结果看可能在G3或其它位置）
    if 'sign_location' in data and data['sign_location']:
        _safe_write_cell(ws, 'G3', data['sign_location'])
    
    # 设置签订日期 - 根据实际模板分析，日期在G5位置（E5:F5是合并单元格）
    if 'sign_date' in data and data['sign_date']:
        sign_date = data['sign_date']
        if isinstance(sign_date, str):
            # 解析字符串日期
            try:
                parsed_date = datetime.strptime(sign_date, '%Y-%m-%d')
                _safe_write_cell(ws, 'F5', parsed_date.strftime('%Y-%m-%d'))  # 写入合并单元格的一部分
            except ValueError:
                _safe_write_cell(ws, 'F5', sign_date)
        elif isinstance(sign_date, datetime):
            _safe_write_cell(ws, 'F5', sign_date.strftime('%Y-%m-%d'))
        else:
            _safe_write_cell(ws, 'F5', sign_date)
    else:
        # 如果没有提供日期，保持原值或置空
        _safe_write_cell(ws, 'F5', data.get('sign_date', ''))
    
    # 设置甲方信息（买方）- 根据实际模板结构
    buyer_info = data.get('buyer', data)  # 兼容直接传递买方信息的情况
    if isinstance(buyer_info, dict):
        if 'name' in buyer_info:
            _safe_write_cell(ws, 'B3', buyer_info['name'])  # 甲方位于B3
        elif 'buyer_name' in buyer_info:
            _safe_write_cell(ws, 'B3', buyer_info['buyer_name'])
    
    # 设置供应商信息（乙方）- 根据实际模板结构，从用户上下文获取信息，没有的话默认置空
    supplier_info = data.get('supplier', {})  # 默认为空字典
    
    # 无论用户是否提供供应商信息，都从用户上下文获取，没提供就置空
    # 优先从supplier子字典中获取，如果没有则从根数据中获取，最终没有则使用空字符串
    supplier_name = supplier_info.get('name') or data.get('supplier_name') or ''
    legal_person = supplier_info.get('legal_person') or data.get('legal_person') or ''
    phone = supplier_info.get('phone') or data.get('phone') or data.get('supplier_phone') or ''
    fax = supplier_info.get('fax') or data.get('fax') or data.get('supplier_fax') or ''
    bank = supplier_info.get('bank') or data.get('bank') or data.get('supplier_bank') or ''
    account = supplier_info.get('account') or data.get('account') or data.get('supplier_account') or ''
    representative = supplier_info.get('representative') or data.get('representative') or data.get('supplier_representative') or ''
    company_name = supplier_info.get('company_name') or data.get('company_name') or data.get('supplier_company_name') or supplier_name or ''
    
    # 根据实际模板分析，供应商信息在右侧区域
    # D24-D30 区域存储供应商详细信息
    _safe_write_cell(ws, 'D24', f'单位名称（盖章）：{company_name}')
    _safe_write_cell(ws, 'D25', f'法定代表人：{legal_person}')
    _safe_write_cell(ws, 'D26', f'委托代表人：{representative}')
    _safe_write_cell(ws, 'D27', f'电    话：{phone}')
    _safe_write_cell(ws, 'D28', f'传    真：{fax}')
    _safe_write_cell(ws, 'D29', f'开户行：{bank}')
    _safe_write_cell(ws, 'D30', f'账    号：{account}')
    
    # 供应商名称可能在B4位置（根据分析结果）
    _safe_write_cell(ws, 'B4', supplier_name)
    
    # 处理商品信息 - 从模板分析看，商品信息从第8行开始
    items = data.get('items', [])
    if isinstance(items, list) and len(items) > 0:
        start_row = 8  # 商品信息从第8行开始
        total_quantity = 0
        total_amount = 0
        
        # 获取原始模板中的商品行范围
        # 首先确定原始模板有多少商品行（通过检查是否有内容）
        original_item_rows = 0
        for row_idx in range(start_row, ws.max_row + 1):
            # 检查B列（品名）是否有内容，如果有则认为是商品行
            if ws[f'B{row_idx}'].value is not None and str(ws[f'B{row_idx}'].value).strip() != '':
                original_item_rows += 1
            else:
                # 如果遇到空行，则停止检查
                break
        
        # 如果新数据比原始数据多，需要插入新行
        if len(items) > original_item_rows:
            for i in range(len(items) - original_item_rows):
                ws.insert_rows(start_row + original_item_rows + i)
        elif len(items) < original_item_rows:
            # 如果新数据比较少，删除多余的行
            for i in range(original_item_rows - len(items)):
                ws.delete_rows(start_row + len(items))
        
        # 填充新的商品数据
        for idx, item in enumerate(items):
            row = start_row + idx
            if isinstance(item, dict):
                # 填充各个字段
                _safe_write_cell(ws, f'A{row}', item.get('serial_number', idx+1))  # 序号
                _safe_write_cell(ws, f'B{row}', item.get('name', item.get('product_name', '')))  # 品名
                _safe_write_cell(ws, f'C{row}', item.get('spec', item.get('specification', '')))  # 规格型号
                quantity = item.get('quantity', item.get('qty', 0))
                _safe_write_cell(ws, f'D{row}', quantity)  # 数量
                _safe_write_cell(ws, f'E{row}', item.get('unit', item.get('measurement_unit', '')))  # 单位
                price = item.get('price', item.get('unit_price', 0))
                _safe_write_cell(ws, f'F{row}', price)  # 单价
                amount = item.get('amount', quantity * price)
                _safe_write_cell(ws, f'G{row}', amount)  # 合计金额
                
                # 累计总数和总金额
                total_quantity += quantity
                total_amount += amount
        
        # 填写总计行 - 更新总计行的数据
        summary_row = start_row + len(items)
        _safe_write_cell(ws, f'D{summary_row}', total_quantity)
        _safe_write_cell(ws, f'G{summary_row}', total_amount)
        
        # 转换总金额为中文大写
        chinese_amount = _number_to_chinese(total_amount)
        # 在适当位置放置中文大写金额
        _safe_write_cell(ws, f'B{summary_row + 1}', f'合计人民币大写：{chinese_amount}')
    
    # 处理交付相关信息
    delivery_info = data.get('delivery', data)
    if isinstance(delivery_info, dict):
        # 更新交货时间
        delivery_date = delivery_info.get('date', delivery_info.get('delivery_date'))
        delivery_days = delivery_info.get('days', delivery_info.get('transport_days', 3))
        
        if delivery_date:
            if isinstance(delivery_date, str):
                try:
                    parsed_date = datetime.strptime(delivery_date, '%Y-%m-%d')
                    delivery_str = f'{parsed_date.year}年{parsed_date.month}月{parsed_date.day}日（另加货运{delivery_days}天）'
                except ValueError:
                    delivery_str = f'{delivery_date}（另加货运{delivery_days}天）'
            elif isinstance(delivery_date, datetime):
                delivery_str = f'{delivery_date.year}年{delivery_date.month}月{delivery_date.day}日（另加货运{delivery_days}天）'
            else:
                delivery_str = f'{delivery_date}（另加货运{delivery_days}天）'
            
            # 根据模板结构，可能需要更新适当的单元格
            _safe_write_cell(ws, 'A14', f'二、 交货时间： {delivery_str}')
    
    # 处理其他合同条款
    if 'quality_requirement' in data:
        _safe_write_cell(ws, 'A16', f'四、 技术标准、质量要求：{data["quality_requirement"]}')
    
    if 'freight_method' in data:
        _safe_write_cell(ws, 'A17', f'五、运费方式及费用承担：{data["freight_method"]}')
    
    if 'payment_method' in data:
        _safe_write_cell(ws, 'A18', f'六、 结算及支付方式：{data["payment_method"]}')
    
    if 'breach_clause' in data:
        _safe_write_cell(ws, 'A19', f'七、违约责任：{data["breach_clause"]}')
    
    # 处理商品信息 - 从items列表
    items = data.get('items', [])
    if isinstance(items, list) and len(items) > 0:
        start_row = 8  # 商品信息从第8行开始
        total_quantity = 0
        total_amount = 0
        
        # 获取原始模板中的商品行范围
        # 首先确定原始模板有多少商品行（通过检查是否有内容）
        original_item_rows = 0
        for row_idx in range(start_row, ws.max_row + 1):
            # 检查B列（品名）是否有内容，如果有则认为是商品行
            if ws[f'B{row_idx}'].value is not None and str(ws[f'B{row_idx}'].value).strip() != '':
                original_item_rows += 1
            else:
                # 如果遇到空行，则停止检查
                break
        
        # 如果新数据比原始数据多，需要插入新行
        if len(items) > original_item_rows:
            for i in range(len(items) - original_item_rows):
                ws.insert_rows(start_row + original_item_rows + i)
        elif len(items) < original_item_rows:
            # 如果新数据比较少，删除多余的行
            for i in range(original_item_rows - len(items)):
                ws.delete_rows(start_row + len(items))
        
        # 填充新的商品数据
        for idx, item in enumerate(items):
            row = start_row + idx
            if isinstance(item, dict):
                # 填充各个字段
                _safe_write_cell(ws, f'B{row}', item.get('name', item.get('product_name', '')))
                _safe_write_cell(ws, f'C{row}', item.get('spec', item.get('specification', '')))
                quantity = item.get('quantity', item.get('qty', 0))
                _safe_write_cell(ws, f'D{row}', quantity)
                _safe_write_cell(ws, f'E{row}', item.get('unit', item.get('measurement_unit', '')))
                price = item.get('price', item.get('unit_price', 0))
                _safe_write_cell(ws, f'F{row}', price)
                amount = item.get('amount', quantity * price)
                _safe_write_cell(ws, f'G{row}', amount)
                _safe_write_cell(ws, f'H{row}', item.get('remark', item.get('note', '')))
                
                # 累计总数和总金额
                total_quantity += quantity
                total_amount += amount
        
        # 填写总计行 - 更新总计行的数据
        summary_row = start_row + len(items)
        _safe_write_cell(ws, f'D{summary_row}', total_quantity)
        _safe_write_cell(ws, f'G{summary_row}', total_amount)
        
        # 转换总金额为中文大写
        chinese_amount = _number_to_chinese(total_amount)
        _safe_write_cell(ws, f'B{summary_row + 1}', chinese_amount)  # 大写金额行通常在总计下方
    
    # 处理交付相关信息 - 现在可以完全根据用户输入动态控制
    delivery_info = data.get('delivery', data)
    if isinstance(delivery_info, dict):
        # 如果用户明确指定需要更新交货时间，则更新A14单元格；否则保持空白让用户填写
        if delivery_info.get('update_delivery_time', False):
            delivery_date = delivery_info.get('date', delivery_info.get('delivery_date'))
            delivery_days = delivery_info.get('days', delivery_info.get('transport_days', 3))
            
            if delivery_date:
                if isinstance(delivery_date, str):
                    try:
                        parsed_date = datetime.strptime(delivery_date, '%Y-%m-%d')
                        delivery_str = f'{parsed_date.year}年{parsed_date.month}月{parsed_date.day}日（另加货运{delivery_days}天）'
                    except ValueError:
                        delivery_str = f'{delivery_date}（另加货运{delivery_days}天）'
                elif isinstance(delivery_date, datetime):
                    delivery_str = f'{delivery_date.year}年{delivery_date.month}月{delivery_date.day}日（另加货运{delivery_days}天）'
                else:
                    delivery_str = f'{delivery_date}（另加货运{delivery_days}天）'
                
                _safe_write_cell(ws, 'A14', f'二、 交货时间： {delivery_str}')
        # 如果没有要求更新交货时间，保持A14单元格空白，让用户填写
        
        # 处理交货地点 - 如果用户明确指定需要更新交货地点，则更新A15；否则保持空白
        if delivery_info.get('update_delivery_location', False):
            delivery_location = delivery_info.get('location', delivery_info.get('delivery_location'))
            if delivery_location:
                _safe_write_cell(ws, 'A15', f'三、 交货地点： {delivery_location}')
        # 如果没有要求更新交货地点，保持A15单元格空白，让用户填写
    
    # 处理其他合同条款
    if 'quality_requirement' in data:
        _safe_write_cell(ws, 'A16', f'四、 技术标准、质量要求：{data["quality_requirement"]}')
    
    if 'freight_method' in data:
        _safe_write_cell(ws, 'A17', f'五、运费方式及费用承担：{data["freight_method"]}')
    
    if 'payment_method' in data:
        _safe_write_cell(ws, 'A18', f'六、 结算及支付方式：{data["payment_method"]}')
    
    if 'breach_clause' in data:
        _safe_write_cell(ws, 'A19', f'七、违约责任：{data["breach_clause"]}')


def _safe_write_cell(worksheet, cell_address, value):
    """
    安全地向单元格写入值，避免合并单元格错误
    """
    try:
        worksheet[cell_address] = value
    except AttributeError as e:
        if 'MergedCell' in str(e):
            # 如果是合并单元格错误，跳过写入
            print(f"跳过合并单元格 {cell_address}: {value}")
        else:
            raise e
    except Exception as e:
        # 对其他错误也进行处理
        print(f"写入单元格 {cell_address} 时出错: {e}")


def _number_to_chinese(num: float) -> str:
    """
    将数字转换为中文大写金额
    
    Args:
        num: 需要转换的数字金额
        
    Returns:
        中文大写金额字符串
    """
    if num == 0:
        return '零元整'
    
    chinese_nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟']
    chinese_big_units = ['', '万', '亿']
    
    num = int(round(abs(num)))  # 取绝对值并转为整数
    result = ''
    unit_index = 0
    
    while num > 0:
        section = num % 10000
        section_str = ''
        section_unit_index = 0
        
        temp = section
        while temp > 0:
            digit = temp % 10
            if digit != 0:
                section_str = chinese_nums[digit] + chinese_units[section_unit_index] + section_str
            elif section_str and not section_str.startswith('零'):
                section_str = '零' + section_str
            temp //= 10
            section_unit_index += 1
        
        if section > 0:
            result = section_str + chinese_big_units[unit_index] + result
        elif result and not result.startswith('零'):
            result = '零' + result
        
        num //= 10000
        unit_index += 1
    
    result = result.rstrip('零')
    return result + '元整'


def _replace_placeholders(ws, data: Dict[str, Any]):
    """
    替换单元格中的占位符
    支持 {{key}} 格式的占位符
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # 替换 {{key}} 格式的占位符
                for key, value in data.items():
                    if isinstance(value, (datetime,)):
                        # 如果是日期时间对象，转换为字符串
                        formatted_value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
                        cell.value = cell.value.replace(f'{{{{{key}}}}}', formatted_value)
                    elif not isinstance(value, (dict, list)):
                        # 不替换嵌套数据结构
                        cell.value = cell.value.replace(f'{{{{{key}}}}}', str(value))


def _process_table_data(ws, data: Dict[str, Any], start_row_marker: Optional[str] = None):
    """
    处理表格数据，支持动态扩展行
    """
    for key, value in data.items():
        if isinstance(value, list) and len(value) > 0:
            # 查找表格起始位置
            table_start_row = None
            
            # 如果提供了特定标记，则查找该标记所在行
            if start_row_marker:
                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    for cell in row:
                        if cell.value and start_row_marker in str(cell.value):
                            table_start_row = row_idx
                            break
                    if table_start_row:
                        break
            else:
                # 默认使用第一个数据行作为表格开始
                table_start_row = 8  # 假设第8行为表格开始
            
            if table_start_row:
                # 复制表头行作为模板
                header_row = list(ws[table_start_row])
                
                # 删除原表格行（保留表头）
                for i in range(len(value)):
                    if i > 0:  # 第一行已经存在，只需插入额外的行
                        ws.insert_rows(table_start_row + i)
                
                # 填充数据
                for idx, item in enumerate(value):
                    row_num = table_start_row + idx
                    for col_idx, cell in enumerate(header_row):
                        col_letter = get_column_letter(col_idx + 1)
                        target_cell = ws[f'{col_letter}{row_num}']
                        
                        # 获取列名（从表头获取）
                        col_name = str(cell.value) if cell.value else f'col_{col_idx}'
                        
                        # 提取列名中的字段名（去掉可能的中文描述）
                        field_key = col_name.lower().strip().replace(' ', '_').replace('（', '_').replace('）', '').replace('(', '_').replace(')', '')
                        
                        # 尝试匹配数据中的键
                        cell_value = None
                        for data_key, data_val in item.items():
                            if field_key == data_key.lower() or field_key.startswith(data_key.lower()):
                                cell_value = data_val
                                break
                        
                        if cell_value is not None:
                            if isinstance(cell_value, datetime):
                                target_cell.value = cell_value.strftime('%Y-%m-%d')
                            else:
                                target_cell.value = cell_value
                
                # 如果有额外的空行，删除它们
                remaining_rows = len(list(ws.iter_rows(min_row=table_start_row + len(value))))
                if remaining_rows > 1:  # 保留一行空行
                    for _ in range(remaining_rows - 1):
                        ws.delete_rows(table_start_row + len(value) + 1)


def fill_template_from_dict(
    template_path: str,
    data: Dict[str, Any],
    output_path: str,
    table_config: Optional[Dict] = None
) -> str:
    """
    从字典数据填充Excel模板
    
    Args:
        template_path: 模板文件路径
        data: 包含填充数据的字典
        output_path: 输出文件路径
        table_config: 表格配置（可选），指定表格的位置和列映射
        
    Returns:
        生成的Excel文件路径
    """
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 检测是否为采购合同模板，并使用专门的处理逻辑
    if _is_purchase_contract_template(ws):
        _process_purchase_contract_template(ws, data)
    else:
        # 递归处理所有数据
        processed_data = _flatten_nested_dict(data)
        
        # 替换单元格中的占位符
        _replace_placeholders(ws, processed_data)
        
        # 处理表格数据
        _process_advanced_table_data(ws, data, table_config)
    
    wb.save(output_path)
    return output_path


def _flatten_nested_dict(d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
    """
    将嵌套字典扁平化，使用点分隔符连接键
    例如: {"supplier": {"name": "ABC"}} -> {"supplier.name": "ABC"}
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_nested_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def _process_advanced_table_data(ws, data: Dict[str, Any], table_config: Optional[Dict] = None):
    """
    处理高级表格数据，支持复杂的数据结构
    """
    for key, value in data.items():
        if isinstance(value, list) and len(value) > 0:
            # 查找对应的表格区域
            table_info = _find_table_by_header(ws, key)
            
            if not table_info and table_config and key in table_config:
                table_info = table_config[key]
            
            if table_info:
                _insert_table_data(ws, value, table_info)


def _find_table_by_header(ws, table_name: str):
    """
    通过表名或标题查找表格位置
    """
    # 在前几行中查找可能的表头
    for row_idx in range(1, 10):  # 检查前10行
        for col_idx in range(1, 20):  # 检查前20列
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and (table_name.lower() in str(cell.value).lower() or 
                              str(cell.value).lower().replace('_', ' ').replace('-', ' ') == table_name.lower()):
                # 找到表头，返回表的起始位置
                return {
                    'start_row': row_idx + 1,  # 数据从下一行开始
                    'start_col': col_idx,
                    'headers': _extract_headers(ws, row_idx, col_idx)
                }
    return None


def _extract_headers(ws, header_row: int, start_col: int):
    """
    从表头行提取列标题
    """
    headers = {}
    col_idx = start_col
    while True:
        cell = ws.cell(row=header_row, column=col_idx)
        if not cell.value or str(cell.value).strip() == '':
            break
        # 使用列字母作为键，方便后续使用
        col_letter = get_column_letter(col_idx)
        headers[col_letter] = str(cell.value).strip()
        col_idx += 1
    return headers


def _insert_table_data(ws, data_list: List[Dict], table_info: Dict):
    """
    向工作表插入表格数据
    """
    start_row = table_info['start_row']
    headers = table_info.get('headers', {})
    
    # 计算需要插入的行数
    rows_needed = len(data_list)
    existing_rows = 1  # 至少有一行已经存在
    
    # 插入所需行数（除了第一行）
    for i in range(rows_needed - existing_rows):
        ws.insert_rows(start_row + existing_rows + i)
    
    # 填充数据
    for idx, row_data in enumerate(data_list):
        current_row = start_row + idx
        
        # 遍历可用的头部列
        for col_letter, header_name in headers.items():
            # 从表头名称中提取字段名
            field_name = header_name.lower().replace(' ', '_').replace('（', '_').replace('）', '').replace('(', '_').replace(')', '')
            
            # 尝试在数据中找到对应的值
            cell_value = None
            for data_key, data_val in row_data.items():
                if field_name == data_key.lower() or field_name.startswith(data_key.lower()) or data_key.lower() in field_name:
                    cell_value = data_val
                    break
            
            if cell_value is not None:
                target_cell = ws[f'{col_letter}{current_row}']
                if isinstance(cell_value, datetime):
                    target_cell.value = cell_value.strftime('%Y-%m-%d')
                else:
                    target_cell.value = cell_value


if __name__ == '__main__':
    # 示例用法
    sample_data = {
        "contract_number": "CG-2025-001",
        "sign_date": "2025-01-15",
        "supplier_name": "示例供应商有限公司",
        "supplier_legal_person": "张三",
        "supplier_phone": "13800138000",
        "supplier_fax": "0731-12345678",
        "supplier_bank": "中国银行示例支行",
        "supplier_account": "1234567890123456789",
        "supplier_representative": "李四",
        "items": [
            {
                "name": "办公用品",
                "spec": "A4纸张",
                "quantity": 100,
                "unit": "包",
                "price": 25,
                "amount": 2500,
                "remark": ""
            },
            {
                "name": "办公用品", 
                "spec": "签字笔",
                "quantity": 200,
                "unit": "支",
                "price": 2,
                "amount": 400,
                "remark": "黑色"
            }
        ],
        "delivery_date": "2025-01-20",
        "quality_requirement": "符合国家标准",
        "freight_method": "乙方负责",
        "payment_method": "货到付款（乙方提供1%增值税普票）",
        "breach_clause": "乙方必须按合同要求时间交货，如未按期交货，予以赔偿延期所致甲方的损失"
    }
    
    # 注意：此示例需要一个实际存在的模板文件
    # output = process_xlsx_template('template.xlsx', sample_data, 'output.xlsx')
    # print(f'Excel文件已生成: {output}')