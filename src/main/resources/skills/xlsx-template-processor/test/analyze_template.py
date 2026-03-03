import pandas as pd
import os
from openpyxl import load_workbook
from pathlib import Path

def analyze_template():
    # 获取模板文件路径
    template_path = Path('E:/study/AI/taorunhui-ai/src/main/resources/skills/xlsx-template-processor/resource/template.xlsx')

    if template_path.exists():
        print('模板文件存在，正在分析...')
        
        # 使用openpyxl加载工作簿以获取更多信息
        wb = load_workbook(template_path)
        print(f'工作表名称: {wb.sheetnames}')
        
        # 分析第一个工作表的结构
        ws = wb.active
        print(f'活动工作表: {ws.title}')
        print(f'最大行数: {ws.max_row}')
        print(f'最大列数: {ws.max_column}')
        
        # 显示前几行的内容以了解结构
        print('\n前10行数据:')
        for row in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for col in range(1, min(8, ws.max_column + 1)):  # 只显示前7列
                cell_value = ws.cell(row=row, column=col).value
                row_values.append(str(cell_value) if cell_value is not None else '')
            print(f'行 {row}: {" | ".join(row_values)}')
            
        # 检查特定区域可能的供应商信息位置
        print('\n供应商相关信息可能的位置:')
        supplier_related_cells = [
            ('B4', '供应商名称'),
            ('D24', '单位名称（盖章）'),
            ('D25', '法定代表人'),
            ('D26', '委托代表人'),
            ('D27', '电话'),
            ('D28', '传真'),
            ('D29', '开户行'),
            ('D30', '账号')
        ]
        
        for cell_addr, desc in supplier_related_cells:
            cell_value = ws[cell_addr].value
            print(f'{cell_addr} ({desc}): {cell_value}')
        
        # 检查是否有其他重要的占位符
        print('\n查找可能的占位符 ({{...}}):')
        placeholder_cells = []
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value and '}}' in cell.value:
                    placeholder_cells.append((f'{cell.column_letter}{cell.row}', cell.value))
        
        if placeholder_cells:
            for addr, value in placeholder_cells[:10]:  # 只显示前10个占位符
                print(f'{addr}: {value}')
        else:
            print('未找到明显的占位符')
    else:
        print('模板文件不存在:', template_path)

if __name__ == "__main__":
    analyze_template()