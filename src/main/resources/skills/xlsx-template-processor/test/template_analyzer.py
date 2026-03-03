"""
模板分析器：使用xlsx技能分析Excel模板结构，为优化提供数据支持
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json


def analyze_template_structure(template_path):
    """
    分析Excel模板的结构
    """
    if not Path(template_path).exists():
        print(f"模板文件不存在: {template_path}")
        return None
    
    print(f'正在分析模板: {template_path}')
    
    # 使用openpyxl加载工作簿
    wb = load_workbook(template_path)
    
    analysis_result = {
        'filename': os.path.basename(template_path),
        'worksheets': [],
        'merged_cells': [],
        'placeholders': [],  # 占位符信息
        'supplier_info_locations': [],  # 供应商信息位置
        'contract_info_locations': []   # 合同信息位置
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        worksheet_info = {
            'name': sheet_name,
            'dimensions': {
                'rows': ws.max_row,
                'columns': ws.max_column
            },
            'sample_data': []
        }
        
        # 收集样本数据（前10行）
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(8, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_info = {
                    'address': f'{get_column_letter(col)}{row}',
                    'value': str(cell.value) if cell.value is not None else '',
                    'is_merged': cell.coordinate in [m.coord for m in ws.merged_cells.ranges]
                }
                row_data.append(cell_info)
            worksheet_info['sample_data'].append(row_data)
        
        # 检测合并单元格
        for merged_range in ws.merged_cells.ranges:
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            analysis_result['merged_cells'].append({
                'range': str(merged_range),
                'top_left_value': str(top_left_cell.value) if top_left_cell.value else ''
            })
        
        # 检测占位符
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if '{{' in cell.value and '}}' in cell.value:
                        analysis_result['placeholders'].append({
                            'address': f'{get_column_letter(col)}{row}',
                            'value': cell.value,
                            'sheet': sheet_name
                        })
        
        # 检测可能的供应商信息位置
        supplier_keywords = ['供应商', '乙方', '单位名称', '法定代表人', '委托代表', '电话', '传真', '开户行', '账号']
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    for keyword in supplier_keywords:
                        if keyword in cell.value:
                            analysis_result['supplier_info_locations'].append({
                                'address': f'{get_column_letter(col)}{row}',
                                'value': cell.value,
                                'keyword': keyword,
                                'sheet': sheet_name
                            })
        
        # 检测可能的合同信息位置
        contract_keywords = ['合同编号', '签约', '签订', '甲方', '买方', '卖方', '日期', '时间']
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    for keyword in contract_keywords:
                        if keyword in cell.value:
                            analysis_result['contract_info_locations'].append({
                                'address': f'{get_column_letter(col)}{row}',
                                'value': cell.value,
                                'keyword': keyword,
                                'sheet': sheet_name
                            })
        
        analysis_result['worksheets'].append(worksheet_info)
    
    return analysis_result


def print_analysis_report(analysis_result):
    """
    打印分析报告
    """
    if not analysis_result:
        return
    
    print(f"\n=== 模板分析报告: {analysis_result['filename']} ===")
    
    # 工作表信息
    print(f"\n工作表数量: {len(analysis_result['worksheets'])}")
    for ws in analysis_result['worksheets']:
        print(f"  - {ws['name']}: {ws['dimensions']['rows']}行 x {ws['dimensions']['columns']}列")
    
    # 合并单元格信息
    print(f"\n合并单元格数量: {len(analysis_result['merged_cells'])}")
    for mc in analysis_result['merged_cells'][:5]:  # 只显示前5个
        print(f"  - {mc['range']}: {mc['top_left_value']}")
    
    # 占位符信息
    print(f"\n占位符数量: {len(analysis_result['placeholders'])}")
    for ph in analysis_result['placeholders'][:10]:  # 只显示前10个
        print(f"  - {ph['sheet']}!{ph['address']}: {ph['value']}")
    
    # 供应商信息位置
    print(f"\n供应商相关字段数量: {len(analysis_result['supplier_info_locations'])}")
    for si in analysis_result['supplier_info_locations']:
        print(f"  - {si['sheet']}!{si['address']}: '{si['value']}' (关键词: {si['keyword']})")
    
    # 合同信息位置
    print(f"\n合同相关字段数量: {len(analysis_result['contract_info_locations'])}")
    for ci in analysis_result['contract_info_locations']:
        print(f"  - {ci['sheet']}!{ci['address']}: '{ci['value']}' (关键词: {ci['keyword']})")


def save_analysis_report(analysis_result, output_path):
    """
    保存分析报告到JSON文件
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(analysis_result, f, ensure_ascii=False, indent=2)
    print(f"\n分析报告已保存至: {output_path}")


def main():
    """
    主函数
    """
    template_path = Path(__file__).parent / 'resource' / 'template.xlsx'
    output_path = Path(__file__).parent / 'template_analysis.json'
    
    # 分析模板结构
    analysis = analyze_template_structure(template_path)
    
    # 打印报告
    print_analysis_report(analysis)
    
    # 保存报告
    if analysis:
        save_analysis_report(analysis, output_path)


if __name__ == "__main__":
    main()